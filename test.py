import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def extract_cell_value(file_path, a, b):
    """
    从 Excel 文件中提取指定单元格的值

    参数:
        file_path (str): Excel 文件路径
        a (int): 行号（从1开始）
        b (int): 列号（从1开始）

    返回:
        cell_value: 指定单元格的值，如果出错则返回 None
    """
    try:
        # 检查行号和列号是否有效
        if a < 1 or b < 1:
            raise ValueError("行号和列号必须从1开始")

        # 读取 Excel 文件
        df = pd.read_excel(file_path)

        # 调整为 0 索引（pandas 使用 0 索引）
        row_index = a - 1
        col_index = b - 1

        # 检查索引是否超出范围
        if row_index >= len(df) or col_index >= len(df.columns):
            raise IndexError("指定的行号或列号超出范围")

        # 提取单元格值
        cell_value = df.iloc[row_index, col_index]
        return cell_value

    except FileNotFoundError:
        print(f"错误：找不到文件 '{file_path}'")
        return None
    except ValueError as ve:
        print(f"值错误：{ve}")
        return None
    except IndexError as ie:
        print(f"索引错误：{ie}")
        return None
    except Exception as e:
        print(f"发生未知错误：{e}")
        return None
def swap_rows_in_excel(file_path, sheet_name, a, b, output_file=None):
    """
    交换 Excel 文件中指定工作表的两行数据，所有单元格数据作为字符串处理

    参数:
        file_path (str): 输入 Excel 文件路径
        sheet_name (str): 工作表名称
        a (int): 第一个行号（从1开始）
        b (int): 第二个行号（从1开始）
        output_file (str, optional): 输出文件路径，默认为原文件名加"_swapped"后缀

    返回:
        bool: 操作成功返回 True，失败返回 False
    """
    try:
        if a < 1 or b < 1 or a == b:
            raise ValueError("行号必须从1开始，且不能相同")

        if not output_file:
            import os
            base, ext = os.path.splitext(file_path)
            output_file = f"{base}_swapped{ext}"

        wb = load_workbook(file_path)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")

        ws = wb[sheet_name]
        max_col = ws.max_column

        # 先读取两行数据并转换为字符串
        row_a_values = [str(ws.cell(row=a, column=i).value) if ws.cell(row=a, column=i).value is not None else ''
                        for i in range(1, max_col + 1)]
        row_b_values = [str(ws.cell(row=b, column=i).value) if ws.cell(row=b, column=i).value is not None else ''
                        for i in range(1, max_col + 1)]

        # 再写入交换后的数据
        for col in range(1, max_col + 1):
            ws.cell(row=a, column=col, value=row_b_values[col - 1])
            ws.cell(row=b, column=col, value=row_a_values[col - 1])

        wb.save(output_file)
        print(f"成功交换第 {a} 行和第 {b} 行，并保存到 {output_file}")
        return True

    except FileNotFoundError:
        print(f"错误：找不到文件 '{file_path}'")
        return False
    except ValueError as ve:
        print(f"值错误：{ve}")
        return False
    except Exception as e:
        print(f"发生未知错误：{e}")
        return False
def swap_columns_in_excel(file_path, sheet_name, a, b, output_file=None):
    """
    交换 Excel 文件中指定工作表的两列数据，所有单元格数据作为字符串处理

    参数:
        file_path (str): 输入 Excel 文件路径
        sheet_name (str): 工作表名称
        a (int): 第一个列号（从1开始）
        b (int): 第二个列号（从1开始）
        output_file (str, optional): 输出文件路径，默认为原文件名加"_swapped"后缀

    返回:
        bool: 操作成功返回 True，失败返回 False
    """
    try:
        if a < 1 or b < 1 or a == b:
            raise ValueError("列号必须从1开始，且不能相同")

        if not output_file:
            import os
            base, ext = os.path.splitext(file_path)
            output_file = f"{base}_swapped{ext}"

        wb = load_workbook(file_path)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")

        ws = wb[sheet_name]
        max_row = ws.max_row

        # 先读取两列数据并转换为字符串
        col_a_values = [str(ws.cell(row=i, column=a).value) if ws.cell(row=i, column=a).value is not None else ''
                        for i in range(1, max_row + 1)]
        col_b_values = [str(ws.cell(row=i, column=b).value) if ws.cell(row=i, column=b).value is not None else ''
                        for i in range(1, max_row + 1)]

        # 再写入交换后的数据
        for row in range(1, max_row + 1):
            ws.cell(row=row, column=a, value=col_b_values[row - 1])
            ws.cell(row=row, column=b, value=col_a_values[row - 1])

        wb.save(output_file)
        print(f"成功交换第 {a} 列和第 {b} 列，并保存到 {output_file}")
        return True

    except FileNotFoundError:
        print(f"错误：找不到文件 '{file_path}'")
        return False
    except ValueError as ve:
        print(f"值错误：{ve}")
        return False
    except Exception as e:
        print(f"发生未知错误：{e}")
        return False
def swap_value(file_path, sheet_name,x1,x2):
    swap_rows_in_excel(file_path, sheet_name,x1+1,x2+1,file_path)
    swap_columns_in_excel(file_path, sheet_name,x1+1,x2+1,file_path)
# 使用示例
if __name__ == "__main__":
    file_path = "2.xlsx"  # 替换为你的文件路径
    for sheet_name in ["2000","2005","2010","2015","2020"]:
        swap_value(file_path,sheet_name,1,4)
        swap_value(file_path,sheet_name,2,3)
        swap_value(file_path,sheet_name,7,9)
        swap_value(file_path,sheet_name,8,9)