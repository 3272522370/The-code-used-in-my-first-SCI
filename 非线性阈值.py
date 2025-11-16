import numpy as np
from scipy.optimize import minimize
import openpyxl
import os
import matplotlib.pyplot as plt
from scipy.interpolate import CubicSpline
import openpyxl
def get_points(input):
    workbook = openpyxl.load_workbook(input)
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    dots=[]
    for row in range(2,sheet.max_row+1):
        x = sheet.cell(row=row,column=3).value
        x=int(x*100)
        y = sheet.cell(row=row,column=4).value
        y = round(y,1)
        dots.append((x,y))
    return dots
def draw_points(points):
    """
    绘制二维点的约束线（凸包边界）
    :param points: 二维点的列表，每个点表示为 (x, y) 元组或列表
    """
    # 将点转换为 numpy 数组
    points = np.array(points)

    # 绘制所有点
    plt.plot(points[:, 0], points[:, 1], 'o')
def divide_points(points,top=1):
    points.sort(key=lambda point: (point[0], point[1]))
    top_points = []
    other_points = []
    for i in range(len(points)):
        if i==len(points)-top:
            top_points.append(points[i])
            break
        if points[i][0]<points[i+top][0]:
            top_points.append(points[i])
        else:
            other_points.append(points[i])
    return top_points,other_points
def polynomial_fit_plot(points, degree=2):
    """
    多项式拟合并绘制拟合曲线
    :param points: 二维点列表，每个点表示为 (x, y) 元组或列表
    :param degree: 多项式的阶数，默认为 3
    """
    # 提取 x 和 y 坐标
    x = np.array([point[0] for point in points])
    y = np.array([point[1] for point in points])

    # 进行多项式拟合
    coefficients = np.polyfit(x, y, degree)
    polynomial = np.poly1d(coefficients)

    # 生成更多的 x 值用于绘制平滑曲线
    x_fit = np.linspace(min(x), max(x), 100)
    y_fit = polynomial(x_fit)

    # 绘制原始数据点和拟合曲线
    plt.scatter(x, y, label='Original Points', color='red')
    plt.plot(x_fit, y_fit, label=f'Polynomial Fit (Degree {degree})', color='blue')
    plt.legend()
    plt.grid(True)
def create_excel(points,path):
    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i in range(len(points)):
        sheet.cell(row=i+1,column=1,value=points[i][0])
        sheet.cell(row=i+1,column=2,value=points[i][1])
    workbook.save(path)
    print(f"成功生成文件{path}")
if __name__=="__main__":
    points = get_points("F:\\水稻和涝渍的关系\\华南（二季晚稻）\\涝渍.xlsx")
    top_points,other_points=divide_points(points)
    print(len(top_points),len(other_points))
    draw_points(other_points)
    polynomial_fit_plot(top_points)
    plt.show()
    #洪涝持续天数(0.9)   洪涝频次(0.9)   平均相对含水量
    create_excel(top_points,r"F:\散点与约束点（水稻）\华南（二季晚稻）\ARWC\constraint_points.xlsx")
    create_excel(other_points,r"F:\散点与约束点（水稻）\华南（二季晚稻）\ARWC\other_points.xlsx")
