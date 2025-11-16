from scipy.stats import pearsonr
import openpyxl
from openpyxl import Workbook
def calculate_correlation_p_value(points):
    """
    计算点坐标列表中横纵坐标之间的皮尔逊相关系数对应的 p 值。

    参数:
    points (list): 点坐标的列表，每个元素是一个包含两个元素的列表或元组，表示一个点的 (x, y) 坐标。

    返回:
    float: 横纵坐标之间的皮尔逊相关系数对应的 p 值。
    """
    # 分离 x 和 y 坐标
    x_coords = [point[1] for point in points]
    y_coords = [point[2] for point in points]

    # 计算皮尔逊相关系数和 p 值
    corr, p_value = pearsonr(x_coords, y_coords)

    return corr,p_value

def get_points(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    points = []
    for row in sheet.rows:
        dry = row[2]
        NPP = row[4]
        TVDI = row[5]
        points.append([dry.value,NPP.value,TVDI.value])
    return points

def cut_points(points,num):
    while points[0][0]==num:
        del points[0]

def part_points(points,num):
    start,end = 1 , 1
    while points[start][0]<num:
        start+=1
    end = start
    while points[end][0]==num:
        end+=1
    return points[start:end]
def dict_to_excel(data_dict, output_path):
    # 创建一个新的工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active

    ws.append(["dry_dura","Pearson","length"])
    # 写入数据
    for key, value in data_dict.items():
        ws.append([key, value[0],value[1]])

    # 保存工作簿
    wb.save(output_path)
def main():
    all_points = get_points("G:\\玉米和旱涝的关系\\merge_excel.xlsx")
    outputs = {}
    for i in range(6,46):
        points = part_points(all_points,i)
        length = len(points)
        if length<=1:
            print(i,length)
            continue
        corr, _ = calculate_correlation_p_value(points)
        outputs[i]=[corr,length]
    dict_to_excel(outputs,"G:\\玉米和旱涝的关系\\干旱Pearson.xlsx")
main()