import openpyxl
import os

def exclude_useless(input,output):
    # 打开工作簿
    workbook = openpyxl.load_workbook(input)

    sheet_names = workbook.sheetnames

    print(workbook.sheetnames)
    for sheet_name in sheet_names:
        print(f'正在处理工作表{sheet_name}')
        # 获取工作表对象
        sheet = workbook[sheet_name]
        index = 1
        de_ls = []
        #存储要删除的行（删完行以后，下面的行会往上补！）
        for row in sheet.rows:
            # print(row)
            cell_NPP = row[3]
            cell_planted = row[4]
            try:
                if cell_NPP.value <= 0 or cell_planted.value!=1:
                    de_ls.append(index)
                else:
                    index+=1
            except:
                de_ls.append(index)#删除有数据空缺的点
        #正式开始删除行
        delete_counts={}
        for idx in de_ls:
            delete_counts[idx]=delete_counts.get(idx,0)+1
        for idx in delete_counts:
            sheet.delete_rows(idx,delete_counts[idx])
    workbook.save(output)

def devide(input,output,V):
    ''' V取值为1代表一级水稻，2代表2级稻'''
    workbook = openpyxl.load_workbook(input)
    print(f'已读取{input}')

    sheet_names = workbook.sheetnames
    for sheet_name in sheet_names:
        print(f'正在处理工作表{sheet_name}')
        sheet = workbook[sheet_name]
        idx = 1
        for row in sheet.rows:
            cell = row[0]
            if cell.value == 2:
                idx += 1
            else:
                break
        if V == 1:
            sheet.delete_rows(1,idx-1)
        elif V == 2:
            sheet.delete_rows(idx,sheet.max_row-idx+1)
    workbook.save(output)

def merge_sheets(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet_names = workbook.sheetnames
    print(f'正在处理文件{file_name}')

    # 创建一个新的工作表用于存储合并后的数据
    if "Combined Sheet" not in sheet_names:
        combined_sheet = workbook.create_sheet(title="Combined Sheet")
    else:
        print('不需要处理')
        return
    #合并工作表
    row_offset = 0
    for sheet_name in sheet_names:
        if sheet_name != "Combined Sheet":
            sheet = workbook[sheet_name]
            for row in sheet.rows:
                new_row = []
                for cell in row:
                    new_row.append(cell.value)
                combined_sheet.append(new_row)
            row_offset += sheet.max_row
    print(f'合并后的工作表的长度为{row_offset}')

    workbook.save(file_name)

def create(file_name):
    workbook = openpyxl.Workbook()
    for i in range(2003,2024):
        workbook.create_sheet(str(i))
    workbook.save(file_name)

def add_column_to_excel(file_path, column_index,output):
    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)
    # 获取工作表
    sheet_name = workbook.sheetnames[0]#只有一个工作表
    print(f"正在处理{sheet_name}")
    worksheet = workbook[sheet_name]

    # 插入列
    worksheet.insert_cols(column_index)

    # 如果提供了列数据，添加数据到列中
    for row in range(1,worksheet.max_row+1):
        dry = worksheet.cell(row=row, column=1)
        wet = worksheet.cell(row=row, column=2)
        value = wet.value-dry.value*8
        worksheet.cell(row=row, column=column_index, value=value)

    # 保存工作簿
    workbook.save(output)

def merge_excel_files(input_paths, output_path):
    # 创建一个新的工作簿
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # 遍历每个输入文件路径
    for input_path in input_paths:
        try:
            # 打开输入文件
            from openpyxl import load_workbook
            print(f"正在处理文件{input_path}")
            input_workbook = load_workbook(input_path)
            input_sheet = input_workbook.active

            # 遍历输入工作表的每一行
            for row in input_sheet.iter_rows(values_only=True):
                # 将行数据添加到新工作表中
                new_sheet.append(row)
        except Exception as e:
            print(f"处理文件 {input_path} 时出现错误: {e}")

    # 保存新的工作簿到输出路径
    try:
        new_workbook.save(output_path)
        print(f"合并完成，结果已保存到 {output_path}")
    except Exception as e:
        print(f"保存文件 {output_path} 时出现错误: {e}")

def main():
    dir = r"F:\水稻和涝渍的关系\华南（二季晚稻）\涝渍数据"
    files = os.listdir(dir)
    for i in range(len(files)):
        files[i] = dir + "\\" + files[i]
    merge_excel_files(files,"test.xlsx")

    # exclude_useless("test.xlsx","test.xlsx")
if __name__=='__main__':
    main()